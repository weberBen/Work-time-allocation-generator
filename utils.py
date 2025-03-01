import pandas as pd
import numpy as np
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

#%%

DIRICHLET_EMPTY_VALUE = 1e-6

#%%

def parse_holiday_dates(year, dates, format="%d/%m"):
    if isinstance(dates, str):
        dates = [dates]
    return [datetime.strptime(date, format).replace(year=year) for date in dates]

def weeks_for_year(year):
    last_week = date(year, 12, 28)
    return last_week.isocalendar().week

def get_week_number(dates):
    """
    Convert a list of dates into week numbers for the specified year.
    """
    if isinstance(dates, str):
        dates = [dates]
    
    week_numbers = [date.isocalendar().week for date in dates]
    return week_numbers

def get_holidays(full_dates, mid_dates):
    full_holidays = get_week_number(full_dates)
    mid_holidays = get_week_number(mid_dates)
    
    full_holidays = {week: full_holidays.count(week) for week in set(full_holidays)}
    mid_holidays = {week: mid_holidays.count(week) for week in set(mid_holidays)}
    return full_holidays, mid_holidays

def adjust_hours_to_target(hours, desired_total):
    """
    Adjust an array of project hours to ensure the sum matches the desired total.
    If the sum exceeds the desired total, hours are reduced proportionally.
    
    Args:
        hours (list or np.array): Array of project hours.
        desired_total (int): The target total sum of hours.
    
    Returns:
        np.array: Adjusted array of project hours.
    """
    hours = np.array(hours)  # Ensure input is a numpy array
    desired_total = round(desired_total, 0)

    # Calculate the excess to be removed
    excess = hours.sum() - desired_total

    if excess == 0:
        return hours
    
    # Calculate proportions for removing hours based on the current distribution
    proportions = hours / hours.sum()
    reduction = np.floor(proportions * excess).astype(int)
    
    # Apply the reduction
    adjusted_hours = hours - reduction
    
    # If the adjustment is slightly off due to rounding, fix it
    while adjusted_hours.sum() != desired_total:
        if adjusted_hours.sum() < desired_total:
            # Remove one hour from the project with the most hours
            idx = adjusted_hours.argmin()
            adjusted_hours[idx] += 1
        
        elif adjusted_hours.sum() > desired_total:
            # Add one hour to the project with the least hours
            idx = adjusted_hours.argmax()
            adjusted_hours[idx] -= 1
    
    return adjusted_hours

#%%

def generate_hours_distribution(dirichlet_params, noise=None, noise_type="uniform", noise_operand="add"):
    """
    Generate a distribution with customizable noise types.
    """
    # Base distribution
    dist_sample = np.random.dirichlet(dirichlet_params)
    if noise is None or noise == 0:
        return dist_sample
    
    # Generate noise based on the specified type
    if noise_type == "uniform":
        noise_factors = np.random.uniform(0, noise, size=dist_sample.shape)
    elif noise_type == "gaussian":
        noise_factors = np.random.normal(loc=0, scale=noise, size=dist_sample.shape)
    elif noise_type == "exponential":
        noise_factors = np.random.exponential(scale=noise, size=dist_sample.shape)
    elif noise_type == "lognormal":
        noise_factors = np.random.lognormal(mean=0, sigma=noise, size=dist_sample.shape)
    elif noise_type == "random":
        noise_factors = np.random.random(dist_sample.shape) * noise
    else:
        raise ValueError(f"Unsupported noise type: {noise_type}")

    noise_factors = np.abs(noise_factors)

    # Apply noise
    if noise_operand == "add":
        altered_dist = dist_sample + noise_factors
    elif noise_operand == "mult":
        altered_dist = dist_sample * (1 + noise_factors)
    else:
        raise ValueError(f"Unsupported noise operand: {noise_type}")
    
    # Preserve zero values from the original
    altered_dist[dist_sample == 0] = 0
    # Renormalize
    altered_dist = altered_dist / altered_dist.sum()

    return altered_dist

        
#%%

def infer_project_distribution(allocation_df):
    """
    Infer the Dirichlet parameters used to generate the project hour distributions.
    
    Args:
        allocation_df (pd.DataFrame): DataFrame where columns are weeks and rows are projects
        
    Returns:
        tuple: (base_distribution, dirichlet_params, estimated_noise)
            - base_distribution: The underlying project distribution proportions
            - dirichlet_params: The estimated Dirichlet parameters
            - estimated_noise: Estimated noise level used in the distribution
    """
    # Convert hours to proportions for each week
    weekly_totals = allocation_df.sum()
    proportions = allocation_df.div(weekly_totals, axis=1)
    
    # Calculate the mean proportion for each project (base distribution)
    base_distribution = proportions.mean(axis=1)
    
    # Calculate the standard deviation of proportions to estimate noise
    prop_std = proportions.std(axis=1)
    estimated_noise = prop_std.mean()
    
    # Reconstruct the Dirichlet parameters
    dirichlet_params = base_distribution
    dirichlet_params[dirichlet_params == 0] = DIRICHLET_EMPTY_VALUE  # Match the original generation logic
    
    return dirichlet_params.values, estimated_noise

#%%

def allocate_hours(
    year, holiday_dates, min_week_hours, max_week_hours, average_week_hours,
    average_rolling_week_hours, tracking_rolling_weeks, project_distribution, max_yearly_overtime, yearly_overtime_variance,
    number_working_days = 5, dirichlet_factor=10, dirichlet_noise=None, dirichlet_noise_type="uniform", dirichlet_noise_operand="add",
    start_week=1, end_week=52, normal_distribution_factor=None
):
    """
    Allocate hours to projects while respecting constraints and variations.
    """
    init_start_week = start_week
    init_end_week = end_week

    project_distribution = np.array(project_distribution)

    weeks = list(range(start_week, end_week + 1))

    full_holidays, mid_holidays = get_holidays(holiday_dates["full"], holiday_dates["mid"])

    reduced_workable_hours = {}
    min_workable_week_hours = [0 for i in range(len(weeks))]
    for week in weeks:
      full_week_holidays = full_holidays.get(week, 0)
      mid_week_holidays = mid_holidays.get(week, 0)

      assert full_week_holidays + mid_week_holidays / 2 <= number_working_days, f"Number of holidays must be less than number of working days in week {week}"

      if normal_distribution_factor is None:
      # Create a normal distribution centered at average_week_hours, then clip to min/max
      # The /3 in is used to set a reasonable standard deviation for the normal distribution
        normal_distribution_factor = (max_week_hours - min_week_hours)/3
      
      week_hours = np.clip(
          np.random.normal(average_week_hours, normal_distribution_factor), 
          min_week_hours, 
          max_week_hours
      )
      
      hours_per_day = week_hours / number_working_days
      reduced_workable_hours[week] = hours_per_day * number_working_days - (full_week_holidays * hours_per_day + mid_week_holidays * hours_per_day / 2)

      min_hours_per_day = min_week_hours / number_working_days
      min_workable_week_hours[week-init_start_week] = min_hours_per_day * number_working_days - (full_week_holidays * min_hours_per_day + mid_week_holidays * min_hours_per_day / 2)


    min_workable_week_hours = np.array(min_workable_week_hours)
    min_workable_week_hours = min_workable_week_hours.round(0)
    min_workable_hours = min_workable_week_hours.sum()

    # Initialize the hours allocation table
    allocation = pd.DataFrame(0, index=[i for i in range(project_distribution.shape[0])], columns=weeks)

    # Process weeks in rolling windows
    for start_week in range(0, len(weeks), tracking_rolling_weeks):
        end_week = min(start_week + tracking_rolling_weeks, len(weeks))
        rolling_weeks = weeks[start_week:end_week]
        
        # Calculate total workable hours for this rolling window
        total_rolling_hours = sum(reduced_workable_hours[w] for w in rolling_weeks)
        
        # Calculate target average hours per week for this window
        target_rolling_hours = len(rolling_weeks) * average_rolling_week_hours

        
        for week in rolling_weeks:
            # Ensure we don't go below minimum hours for this week
            reduced_workable_hours[week] = max(reduced_workable_hours[week], min_workable_week_hours[week-init_start_week])
                
        # Scale hours if needed to meet rolling window average constraint and minimum hours
        while total_rolling_hours > target_rolling_hours:
            scale_factor = target_rolling_hours / total_rolling_hours
            
            for week in rolling_weeks:
                # Calculate scaled hoursstart_week
                scaled_hours = reduced_workable_hours[week] * scale_factor
                # Ensure we don't go below minimum hours for this week
                reduced_workable_hours[week] = max(scaled_hours, min_workable_week_hours[week-init_start_week])
            # Recalculate total rolling hours after scaling
            total_rolling_hours = sum(reduced_workable_hours[w] for w in rolling_weeks)
        
        # For each week in the rolling window, distribute hours across projects
        for week in rolling_weeks:
            week_hours = reduced_workable_hours[week]
            # Use Dirichlet distribution to randomly allocate hours while maintaining proportions

            dirichlet_params = project_distribution * dirichlet_factor
            dirichlet_params[dirichlet_params == 0] = DIRICHLET_EMPTY_VALUE

            project_hours = generate_hours_distribution(dirichlet_params,
                                                        noise=dirichlet_noise,
                                                        noise_type=dirichlet_noise_type,
                                                        noise_operand=dirichlet_noise_operand,
                                                    )
            project_hours *= week_hours
            project_hours = np.round(project_hours, 0)
            
            project_hours = adjust_hours_to_target(project_hours, week_hours)
            
            allocation.loc[:, week] = project_hours
        
        weekly_totals = allocation.iloc[:, start_week:end_week].sum(axis=0)
        if weekly_totals.mean() > average_rolling_week_hours:
            # While average is still above target
            while weekly_totals.mean() > average_rolling_week_hours:
                # Find week with highest hours that's still above minimum
                valid_weeks = weekly_totals[weekly_totals > min_workable_week_hours[weekly_totals.index-init_start_week]]
                if len(valid_weeks) == 0:
                    break
                
                # Take random index in valid weeks to avoid bias and give more natural distribution
                idx_week = valid_weeks.index[np.random.randint(len(valid_weeks))]
                
                # Reduce highest week by 1 hour while maintaining minimums
                allocation.loc[:, idx_week] = adjust_hours_to_target(
                    allocation.loc[:, idx_week].to_list(), 
                    weekly_totals[idx_week] - 1
                )
                weekly_totals[idx_week] -= 1

    # Calculate total hours exceeding minimum workable hours
    number_adjustments = 0
    yearly_overtime = allocation.sum().sum() - min_workable_hours
    if yearly_overtime_variance == 0:
      allowed_yearly_overtime = max_yearly_overtime
    else:
      allowed_yearly_overtime = np.random.randint(max(0, max_yearly_overtime - yearly_overtime_variance), max_yearly_overtime)

    if yearly_overtime > allowed_yearly_overtime:
      # Get weekly totals and minimum requirements
      weekly_totals = allocation.sum()
      min_workable_week_hours_series = pd.Series(min_workable_week_hours, index=weeks)

      # While we're over the max yearly hours
      while yearly_overtime > allowed_yearly_overtime:
        # Find weeks that are still above their minimum
        valid_weeks = weekly_totals[weekly_totals > min_workable_week_hours_series]
        if len(valid_weeks) == 0:
          break

        # Take random index in valid weeks to avoid bias and give more natural distribution
        idx_week = valid_weeks.index[np.random.randint(len(valid_weeks))]
        
        # Reduce by 1 hour
        allocation.loc[:, idx_week] = adjust_hours_to_target(allocation.loc[:, idx_week].to_list(), allocation.loc[:, idx_week].sum() - 1)
        
        weekly_totals[idx_week] -= 1
        yearly_overtime -= 1
        number_adjustments += 1
        
    return allocation

#%%

def verify_allocation_constraints(allocation_df, year, holiday_dates, min_week_hours, max_week_hours, 
                                average_rolling_week_hours, tracking_rolling_weeks,
                                average_week_hours, max_yearly_overtime, yearly_overtime_variance,
                                project_distribution, dirichlet_factor, dirichlet_noise,
                                number_working_days=5, start_week=1, end_week=52):
    """
    Verify that the allocation meets all constraints:
    - Weekly hours within min/max bounds
    - Rolling average under limit 
    - Overall average close to target
    
    Args:
        allocation_df: DataFrame with weeks as columns and projects as rows
        min_week_hours: Minimum allowed hours per week
        max_week_hours: Maximum allowed hours per week
        average_rolling_week_hours: Maximum rolling average hours
        tracking_rolling_weeks: Number of weeks for rolling average
        average_week_hours: Target average hours per week
        tolerance: Allowed deviation from average_week_hours
        
    Returns:
        bool: True if all constraints are met, False otherwise
    """
    # Get weekly totals
    weekly_totals = allocation_df.sum()

    weeks = list(range(start_week, end_week + 1))

    full_holidays, mid_holidays = get_holidays(holiday_dates["full"], holiday_dates["mid"])

    # Calculate reduced workable hours for each week
    min_weeks = [0 for i in range(len(weeks))]
    for idx, week in enumerate(weeks):
      full_week_holidays = full_holidays.get(week, 0)
      mid_week_holidays = mid_holidays.get(week, 0)

      hours_per_day = min_week_hours / number_working_days
      min_weeks[idx] = hours_per_day * number_working_days - (full_week_holidays * hours_per_day + mid_week_holidays * hours_per_day / 2)
    
    min_weeks = np.array(min_weeks).round(0)
        
    # Check min/max weekly constraints, using reduced_workable_hours as minimum
    for week_idx, (total, min_week) in enumerate(zip(weekly_totals, min_weeks)):
        if total < min_week:
            print(f"Week {week_idx + 1} violates minimum hours constraint: {total:.1f} < {min_week:.1f}")
            return False
        if total > max_week_hours:
            print(f"Week {week_idx + 1} violates maximum hours constraint: {total:.1f} > {max_week_hours:.1f}")
            return False
    
    for start_week in range(0, len(weeks), tracking_rolling_weeks):
        end_week = min(start_week + tracking_rolling_weeks, len(weeks))
        rolling_window = weekly_totals[start_week:end_week]

        print("Checking rolling weeks", start_week+1, end_week)

        rolling_avg = rolling_window.mean()
        if rolling_avg > average_rolling_week_hours:
            print(f"\tRolling average exceeds limit in weeks {start_week+1}-{end_week+1}", "Computed average", rolling_avg, "Max allowed", average_rolling_week_hours)
            return False
        else:
            print(f"\tRolling average :", rolling_avg, "Max allowed", average_rolling_week_hours)
                
    # Check overall average
    total_avg = weekly_totals.mean()
    total_med = weekly_totals.median()
    print("")
    print(f"Overall week hours average :", total_avg, "Target average :", average_week_hours)
    print(f"Overall week hours median :", total_med, "Target average :", average_week_hours)
    
    # Calculate quartiles
    q1, q3 = np.percentile(weekly_totals, [25, 75])
    print("")
    print(f"First quartile (Q1) of week hours:", q1)
    print(f"Third quartile (Q3) of week hours:", q3)
    print(f"Interquartile range (IQR):", q3 - q1)

    
    yearly_overtime_hours = weekly_totals.sum() - sum(min_weeks)
    print("")
    print(f"Yearly overtime hours :", yearly_overtime_hours, "Max allowed :", max_yearly_overtime, "Variance :", yearly_overtime_variance)

    if yearly_overtime_hours < 0:
        print("Yearly overtime hours is negative", yearly_overtime_hours)
        return False
    
    if yearly_overtime_hours > max_yearly_overtime:
      print("Exceeding yearly hours", yearly_overtime_hours - max_yearly_overtime)
      return False
    
    estimated_project_distribution, estimated_noise = infer_project_distribution(allocation_df)
    project_distribution_diff = np.abs(estimated_project_distribution - project_distribution)
    project_distribution_mae = np.mean(project_distribution_diff)
    
    print("")
    print("Real project distribution:", project_distribution)
    print("Estimated project distribution:", estimated_project_distribution)
    print("Project distribution diff:", project_distribution_mae)

    if project_distribution_mae > 0.1:
      print(f"WARNING: Project distribution from data is not close enough to the target distribution (diff: {project_distribution_mae})")

    return True

#%%

def to_excel_file(allocation_df, project_names,
                  title="Time allocation (in hours)",
                  week_col_prefix="W",
                  project_col_title="Projects",
                  total_col_title="Total",
                  file_prefix=None,
                  display_all_weeks=True):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Display week over the entire year even if the desired period is truncated
    if display_all_weeks:
        allocation_df = allocation_df.copy()
        num_week = 52
        min_week = min(allocation_df.columns)
        max_week = max(allocation_df.columns)

        # Add missing weeks at start
        for week in range(1, min_week):
            allocation_df[week] = 0

        # Add missing weeks at end
        for week in range(max_week + 1, num_week + 1):
            allocation_df[week] = 0

        # Sort columns by week number
        allocation_df = allocation_df.reindex(sorted(allocation_df.columns), axis=1)

    # Write the header
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(allocation_df.columns) + 2)

    ws.column_dimensions['A'].width = max([len(t) for t in project_names]) + 2

    # Write week numbers as column headers starting from B2
    for col, week in enumerate(allocation_df.columns, start=2):
        ws.cell(row=2, column=col, value=f"{week_col_prefix}{week}")
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=2, column=1, value=project_col_title)
    ws.cell(row=2, column=len(allocation_df.columns) + 2, value=total_col_title)

    # Style the header row
    for col in range(1, len(allocation_df.columns) + 2):
        ws.cell(row=2, column=col).font = Font(bold=False, italic=True)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
    # row totals
    col = len(allocation_df.columns) + 2
    ws.cell(row=2, column=col).font = Font(bold=True, italic=False)
    ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Write project names and values
    for row, project in enumerate(allocation_df.index, start=3):
        ws.cell(row=row, column=1, value=project_names[row - 3])
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=row, column=1).font = Font(bold=False, italic=True)
        for col, value in enumerate(allocation_df.loc[project], start=2):
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        # Add row total formula
        last_col = get_column_letter(len(allocation_df.columns) + 1)
        ws.cell(row=row, column=len(allocation_df.columns) + 2, 
                value=f"=SUM(B{row}:{last_col}{row})")
        ws.cell(row=row, column=len(allocation_df.columns) + 2).alignment = Alignment(horizontal="center", vertical="center")

    # Add "Total" row
    total_row = len(allocation_df.index) + 3
    ws.cell(row=total_row, column=1, value=total_col_title)
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Add column total formulas
    for col in range(2, len(allocation_df.columns) + 2):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({col_letter}3:{col_letter}{total_row - 1})")
        ws.cell(row=total_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Add grand total formula - sum of row totals
    ws.cell(row=total_row, column=len(allocation_df.columns) + 2,
            value=f"=SUM({get_column_letter(len(allocation_df.columns) + 2)}3:{get_column_letter(len(allocation_df.columns) + 2)}{total_row - 1})")
    ws.cell(row=total_row, column=len(allocation_df.columns) + 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=len(allocation_df.columns) + 2).font = Font(bold=True)

    # Apply grey fill to the total row and column
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col in range(1, len(allocation_df.columns) + 3):
        ws.cell(row=total_row, column=col).fill = grey_fill
    for row in range(2, total_row + 1):
        ws.cell(row=row, column=len(allocation_df.columns) + 2).fill = grey_fill

    # Add grid lines (borders) to all cells
    thin_border = Border(left=Side(style="thin"),
                         right=Side(style="thin"),
                         top=Side(style="thin"),
                         bottom=Side(style="thin"))
    for row in range(2, total_row + 1):
        for col in range(1, len(allocation_df.columns) + 3):
            ws.cell(row=row, column=col).border = thin_border

    
    # # Adjust column width
    # for col in range(1, len(allocation_df.columns) + 3):
    #     ws.column_dimensions[get_column_letter(col)].width = 10

    # Save the workbook
    filename = f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    if file_prefix is not None:
        filename = f"{file_prefix}{filename}"
    wb.save(filename)

    return filename
